from openpyxl import load_workbook


def load_metadata(file_path="metadata.xlsx"):
    workbook = load_workbook(file_path)
    sheet = workbook.active

    videos = []

    headers = [cell.value for cell in sheet[1]]

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue

        videos.append(dict(zip(headers, row)))

    return videos


if __name__ == "__main__":
    data = load_metadata()

    for item in data:
        print(item)
